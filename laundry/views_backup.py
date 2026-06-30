from django.shortcuts import ren
# Custom decorator for owner-only access
def owner_required(view_func):
    from django.contrib.auth.decorators import login_required, user_passes_test
    decorated_func = login_required(view_func)
    return user_passes_test(
        lambda u: u.is_superuser or u.is_staff,
        login_url='/login/',
        redirect_field_name=None
    )(decorated_func)der, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout
from django.db.models import Sum
from django.utils import timezone
from datetime import timedelta
from django.http import JsonResponse, HttpResponse
import json
from decimal import Decimal
import urllib.parse

from .models import (
    Order, Customer, Expense, Journal, LedgerAccount, JournalEntry, 
    Asset, Liability, Equity, Transaction
)
from .forms import OrderForm, CustomerForm, ExpenseForm

# Smart Transaction
from .transaction_analyzer import TransactionAnalyzer
from .accounting_generator import AccountingEntryGenerator


# ============================================================
# DASHBOARD
# ============================================================
@login_required
def dashboard(request):
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    
    hour = timezone.now().hour
    if hour < 12:
        greeting = "morning"
    elif hour < 17:
        greeting = "afternoon"
    else:
        greeting = "evening"
    
    received_orders = Order.objects.filter(status='received').count()
    washing_orders = Order.objects.filter(status='washing').count()
    drying_orders = Order.objects.filter(status='drying').count()
    folding_orders = Order.objects.filter(status='folding').count()
    ready_orders = Order.objects.filter(status='ready').count()
    completed_orders = Order.objects.filter(status='completed').count()
    
    progress_orders = washing_orders + drying_orders + folding_orders
    
    paid_orders = Order.objects.filter(payment_status='paid')
    unpaid_orders = Order.objects.filter(payment_status='unpaid')
    partial_orders = Order.objects.filter(payment_status='partial')
    
    paid_count = paid_orders.count()
    unpaid_count = unpaid_orders.count()
    partial_count = partial_orders.count()
    
    total_revenue = paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    pending_revenue = unpaid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    pending_orders_count = unpaid_orders.count() + partial_orders.count()
    
    today_orders = Order.objects.filter(created_at__date=today)
    today_paid_orders = today_orders.filter(payment_status='paid')
    today_revenue = today_paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    today_expenses = Expense.objects.filter(date=today).aggregate(Sum('amount'))['amount__sum'] or 0
    
    weekly_orders = Order.objects.filter(created_at__date__gte=week_ago)
    weekly_orders_count = weekly_orders.count()
    weekly_paid_orders = weekly_orders.filter(payment_status='paid')
    weekly_revenue = weekly_paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    week_data = []
    week_labels = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        day_orders = Order.objects.filter(created_at__date=day, payment_status='paid')
        revenue = day_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        week_data.append(float(revenue))
        week_labels.append(day.strftime('%a'))
    
    last_week_start = week_ago - timedelta(days=7)
    last_week_orders = Order.objects.filter(
        created_at__date__gte=last_week_start,
        created_at__date__lt=week_ago,
        payment_status='paid'
    )
    last_week_revenue = last_week_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    if last_week_revenue > 0:
        weekly_percentage = round((weekly_revenue - last_week_revenue) / last_week_revenue * 100)
    else:
        weekly_percentage = 100 if weekly_revenue > 0 else 0
    
    month_ago = today - timedelta(days=30)
    new_customers = Customer.objects.filter(created_at__date__gte=month_ago).count()
    
    total_orders_count = Order.objects.count()
    if total_orders_count > 0:
        completion_rate = round((completed_orders / total_orders_count) * 100)
    else:
        completion_rate = 0
    
    ready_orders_list = Order.objects.filter(status='ready').order_by('-created_at')[:10]
    recent_orders = Order.objects.all().order_by('-created_at')[:10]
    total_customers = Customer.objects.count()
    
    context = {
        'greeting': greeting,
        'received_orders': received_orders,
        'washing_orders': washing_orders,
        'drying_orders': drying_orders,
        'folding_orders': folding_orders,
        'progress_orders': progress_orders,
        'ready_orders': ready_orders,
        'completed_orders': completed_orders,
        'paid_count': paid_count,
        'unpaid_count': unpaid_count,
        'partial_count': partial_count,
        'total_revenue': total_revenue,
        'pending_revenue': pending_revenue,
        'pending_orders_count': pending_orders_count,
        'today_orders': today_orders.count(),
        'today_revenue': today_revenue,
        'today_expenses': today_expenses,
        'today_profit': today_revenue - today_expenses,
        'weekly_orders': weekly_orders_count,
        'weekly_revenue': weekly_revenue,
        'weekly_percentage': weekly_percentage,
        'week_data': week_data,
        'week_labels': week_labels,
        'total_customers': total_customers,
        'new_customers': new_customers,
        'completion_rate': completion_rate,
        'ready_orders_list': ready_orders_list,
        'recent_orders': recent_orders,
    }
    return render(request, 'laundry/dashboard.html', context)


# ============================================================
# ORDERS
# ============================================================
@login_required
def order_list(request):
    orders = Order.objects.all().order_by('-created_at')
    return render(request, 'laundry/orders.html', {'orders': orders})

@login_required
def order_create(request):
    if request.method == 'POST':
        phone = request.POST.get('customer_phone')
        name = request.POST.get('customer_name')
        location = request.POST.get('customer_location', '')
        apartment_name = request.POST.get('customer_apartment', '')
        floor = request.POST.get('customer_floor', '')
        door_number = request.POST.get('customer_door', '')
        existing_customer_id = request.POST.get('existing_customer_id')
        
        if existing_customer_id:
            customer = get_object_or_404(Customer, id=existing_customer_id)
        else:
            customer = Customer.objects.create(
                name=name,
                phone=phone,
                location=location,
                apartment_name=apartment_name,
                floor=floor,
                door_number=door_number
            )
        
        try:
            weight_kg = float(request.POST.get('weight_kg', 0))
            price_per_kg = float(request.POST.get('price_per_kg', 100))
            paid_amount = float(request.POST.get('paid_amount', 0))
        except ValueError:
            weight_kg = 0
            price_per_kg = 100
            paid_amount = 0
        
        order = Order(
            customer=customer,
            items_description=request.POST.get('items_description', ''),
            weight_kg=weight_kg,
            price_per_kg=price_per_kg,
            status=request.POST.get('status', 'received'),
            payment_status=request.POST.get('payment_status', 'unpaid'),
            paid_amount=paid_amount,
            notes=request.POST.get('notes', ''),
            created_by=request.user
        )
        order.save()  # This will automatically create accounting entries via the Order model's save method
        
        return redirect('order_list')
    
    return render(request, 'laundry/order_form.html', {'title': 'New Order'})
@login_required
def order_edit(request, pk):
    order = get_object_or_404(Order, pk=pk)
    if request.method == 'POST':
        form = OrderForm(request.POST, instance=order)
        if form.is_valid():
            form.save()
            return redirect('order_list')
    else:
        form = OrderForm(instance=order)
    return render(request, 'laundry/order_form.html', {'form': form, 'title': 'Edit Order'})

@login_required
def update_order_status(request, order_id, status):
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, id=order_id)
            order.status = status
            order.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required
def update_order_payment(request, order_id, payment_status):
    if request.method == 'POST':
        try:
            order = get_object_or_404(Order, id=order_id)
            order.payment_status = payment_status
            if payment_status == 'paid':
                order.paid_amount = order.total_amount
            order.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required
def mark_all_ready_completed(request):
    if request.method == 'POST':
        try:
            count = Order.objects.filter(status='ready').update(status='completed')
            return JsonResponse({'success': True, 'message': f'{count} order(s) marked as completed!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


# ============================================================
# CUSTOMERS
# ============================================================
@login_required
def customer_list(request):
    customers = Customer.objects.all().order_by('-total_orders')
    return render(request, 'laundry/customers.html', {'customers': customers})

@login_required
def customer_create(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm()
    return render(request, 'laundry/customer_form.html', {'form': form, 'title': 'New Customer'})

@login_required
def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    if request.method == 'POST':
        form = CustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return redirect('customer_list')
    else:
        form = CustomerForm(instance=customer)
    return render(request, 'laundry/customer_form.html', {'form': form, 'title': 'Edit Customer'})

@login_required
def find_customer_by_phone(request, phone):
    try:
        customer = Customer.objects.get(phone=phone)
        return JsonResponse({
            'exists': True,
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone,
            'location': customer.location or '',
            'apartment_name': customer.apartment_name or '',
            'floor': customer.floor or '',
            'door_number': customer.door_number or ''
        })
    except Customer.DoesNotExist:
        return JsonResponse({'exists': False})


# ============================================================
# EXPENSES
# ============================================================
@login_required
def expense_list(request):
    expenses = Expense.objects.all().order_by('-date')
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    
    assets = Asset.objects.all().order_by('-created_at')
    total_assets_value = assets.aggregate(Sum('current_value'))['current_value__sum'] or 0
    
    category_choices = Expense.CATEGORY_CHOICES
    
    context = {
        'expenses': expenses,
        'total_expenses': total_expenses,
        'expenses_count': expenses.count(),
        'assets': assets,
        'assets_count': assets.count(),
        'total_assets_value': total_assets_value,
        'category_choices': category_choices,
    }
    return render(request, 'laundry/expenses.html', context)
@login_required
def expense_create(request):
    if request.method == 'POST':
        description = request.POST.get('description')
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        notes = request.POST.get('notes', '')
        entry_type = request.POST.get('entry_type', 'expense')
        
        if not description or not amount:
            return redirect('expense_list')
        
        try:
            if entry_type == 'asset':
                category = 'asset_purchase'
            
            expense = Expense.objects.create(
                description=description,
                category=category or 'other',
                amount=int(amount),
                notes=notes,
                created_by=request.user
            )
            # The Expense model's save method will automatically create accounting entries
            return redirect('expense_list')
        except Exception as e:
            return redirect('expense_list')
    
    return render(request, 'laundry/expense_form.html', {'title': 'Add Expense'})

@login_required
def expense_edit(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    
    if request.method == 'POST':
        description = request.POST.get('description')
        category = request.POST.get('category')
        amount = request.POST.get('amount')
        notes = request.POST.get('notes', '')
        entry_type = request.POST.get('entry_type', 'expense')
        
        if not description or not amount:
            return redirect('expense_list')
        
        try:
            if entry_type == 'asset':
                category = 'asset_purchase'
            
            expense.description = description
            expense.category = category or 'other'
            expense.amount = int(amount)
            expense.notes = notes
            expense.save()
            return redirect('expense_list')
        except Exception as e:
            return redirect('expense_list')
    
    return render(request, 'laundry/expense_form.html', {
        'expense': expense,
        'title': 'Edit Expense',
        'is_edit': True
    })

@login_required
def expense_delete(request, pk):
    if request.method == 'POST':
        try:
            expense = get_object_or_404(Expense, pk=pk)
            expense.delete()
            return JsonResponse({'success': True, 'message': 'Expense deleted successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# ============================================================
# REPORTS & RECEIPTS
# ============================================================
@login_required
def reports(request):
    today = timezone.now().date()
    week_ago = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)
    
    total_orders_count = Order.objects.count()
    total_revenue = Order.objects.filter(payment_status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_expenses = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    total_customers = Customer.objects.count()
    net_profit = total_revenue - total_expenses
    
    weekly_revenue = Order.objects.filter(created_at__date__gte=week_ago, payment_status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    monthly_revenue = Order.objects.filter(created_at__date__gte=month_ago, payment_status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Calculate daily data for chart
    daily_data = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        revenue = Order.objects.filter(created_at__date=day, payment_status='paid').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        expenses = Expense.objects.filter(date=day).aggregate(Sum('amount'))['amount__sum'] or 0
        daily_data.append({
            'date': day.strftime('%Y-%m-%d'),
            'revenue': float(revenue),
            'expenses': float(expenses),
            'profit': float(revenue - expenses)
        })
    
    context = {
        'total_orders': total_orders_count,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'total_profit': net_profit,
        'total_customers': total_customers,
        'weekly_revenue': weekly_revenue,
        'monthly_revenue': monthly_revenue,
        'daily_data': daily_data,
    }
    return render(request, 'laundry/reports.html', context)
@login_required
def receipt(request, order_id):
    order = get_object_or_404(Order, id=order_id)
    return render(request, 'laundry/receipt.html', {'order': order})

@login_required
def receipt_data(request, order_id):
    """API endpoint to get receipt data for WhatsApp sharing"""
    order = get_object_or_404(Order, id=order_id)
    data = {
        'success': True,
        'order_number': order.order_number,
        'date': order.created_at.strftime('%Y-%m-%d %H:%M'),
        'customer_name': order.customer.name,
        'customer_phone': order.customer.phone,
        'customer_location': order.customer.location or '',
        'weight_kg': order.weight_kg,
        'price_per_kg': order.price_per_kg,
        'total_amount': order.total_amount,
        'paid_amount': order.paid_amount,
        'balance': order.remaining_balance,
        'status': order.get_status_display(),
    }
    return JsonResponse(data)

@login_required
def receipt_pdf(request, order_id):
    """Download receipt as PDF"""
    order = get_object_or_404(Order, id=order_id)
    from .pdf_generator import generate_pdf_receipt
    
    pdf_buffer = generate_pdf_receipt(order)
    
    response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="receipt_{order.order_number}.pdf"'
    return response

@login_required
def receipt_whatsapp(request, order_id):
    """Send receipt as text message via WhatsApp"""
    order = get_object_or_404(Order, id=order_id)
    
    # Get customer phone
    phone = order.customer.phone.replace(' ', '').replace('-', '')
    if phone.startswith('0'):
        phone = '254' + phone[1:]
    if not phone.startswith('254'):
        phone = '254' + phone
    
    # Build the WhatsApp message - NO EMOJIS
    message = f"""ZENCLEAN LAUNDRY 


Order #: {order.order_number}
Date: {order.created_at.strftime('%Y-%m-%d %H:%M')}

CUSTOMER DETAILS:
Name: {order.customer.name}
Phone: {order.customer.phone}
Location: {order.customer.location or 'N/A'}

ORDER DETAILS:
Items: {order.items_description or 'N/A'}
Weight: {order.weight_kg} kg
Rate: KSh {order.price_per_kg}/kg

PAYMENT SUMMARY:
Total: KSh {order.total_amount:,}
Paid: KSh {order.paid_amount:,}"""
    
    if order.remaining_balance > 0:
        message += f"\nBalance: KSh {order.remaining_balance:,}"
    else:
        message += "\nFully Paid"
    
    message += f"""

Status: {order.get_status_display()}

Thank you for choosing ZenClean!
Ready in 24 hours

ZenClean Laundry
0712345678
Westlands, Nairobi"""
    
    # URL encode the message
    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    
    whatsapp_url = f"https://wa.me/{phone}?text={encoded_message}"
    
    return JsonResponse({
        'success': True,
        'whatsapp_url': whatsapp_url,
        'phone': phone,
        'message': message
    })
# ============================================================
# ACCOUNTING - FULL SYSTEM
# ============================================================
@login_required
def accounting_dashboard(request):
    from django.db.models import Sum
    from .models import LedgerAccount, JournalEntry, Journal, Asset, Liability, Equity, Order, Expense
    
    # ----- STEP 1: Calculate from SOURCE MODELS (not just journal entries) -----
    
    # 1. Revenue from PAID orders
    paid_orders = Order.objects.filter(payment_status='paid')
    total_revenue = paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # 2. Expenses from Expense model
    total_expenses = Expense.objects.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # 3. Assets from Asset model
    assets_list = Asset.objects.all().order_by('asset_code')
    total_assets = assets_list.aggregate(Sum('current_value'))['current_value__sum'] or 0
    
    # 4. Liabilities from Liability model
    liabilities_list = Liability.objects.all().order_by('liability_code')
    total_liabilities = liabilities_list.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # 5. Equity from Equity model
    equity_list = Equity.objects.all().order_by('equity_code')
    total_equity = equity_list.aggregate(Sum('amount'))['amount__sum'] or 0
    
    # 6. Net Profit/Loss
    net_profit = total_revenue - total_expenses
    
    # ----- STEP 2: Get journals with entries -----
    journals = Journal.objects.filter(entries__isnull=False).distinct().order_by('-created_at')[:50]
    
    # ----- STEP 3: Prepare T-account data from journal entries -----
    # Get all accounts that have journal entries
    accounts_with_entries = []
    all_accounts = LedgerAccount.objects.filter(is_active=True).order_by('account_code')
    
    for account in all_accounts:
        has_entries = JournalEntry.objects.filter(account=account).exists()
        if has_entries:
            accounts_with_entries.append(account)
    
    # Also include accounts that represent assets/liabilities/equity from models
    # Add Fixed Assets account if there are assets
    if total_assets > 0:
        try:
            fixed_account, _ = LedgerAccount.objects.get_or_create(
                account_code='1400',
                defaults={
                    'name': 'Fixed Assets',
                    'account_type': 'asset',
                    'normal_balance': 'D',
                    'opening_balance': 0,
                    'current_balance': 0,
                    'is_active': True
                }
            )
            fixed_account.current_balance = total_assets
            fixed_account.save()
            if fixed_account not in accounts_with_entries:
                accounts_with_entries.append(fixed_account)
        except:
            pass
    
    # Add Revenue account
    try:
        rev_account, _ = LedgerAccount.objects.get_or_create(
            account_code='4000',
            defaults={
                'name': 'Laundry Service Revenue',
                'account_type': 'revenue',
                'normal_balance': 'C',
                'opening_balance': 0,
                'current_balance': 0,
                'is_active': True
            }
        )
        rev_account.current_balance = total_revenue
        rev_account.save()
        if rev_account not in accounts_with_entries:
            accounts_with_entries.append(rev_account)
    except:
        pass
    
    # Add Expense account
    try:
        exp_account, _ = LedgerAccount.objects.get_or_create(
            account_code='5900',
            defaults={
                'name': 'Total Expenses',
                'account_type': 'expense',
                'normal_balance': 'D',
                'opening_balance': 0,
                'current_balance': 0,
                'is_active': True
            }
        )
        exp_account.current_balance = total_expenses
        exp_account.save()
        if exp_account not in accounts_with_entries:
            accounts_with_entries.append(exp_account)
    except:
        pass
    
    # ----- STEP 4: Prepare T-account data -----
    t_account_data = []
    for account in accounts_with_entries:
        entries = JournalEntry.objects.filter(account=account).order_by('-created_at')
        debit_entries = entries.filter(debit__gt=0)[:20]
        credit_entries = entries.filter(credit__gt=0)[:20]
        total_debit = entries.aggregate(Sum('debit'))['debit__sum'] or 0
        total_credit = entries.aggregate(Sum('credit'))['credit__sum'] or 0
        
        # Calculate balance
        if account.normal_balance == 'D':
            balance = account.opening_balance + total_debit - total_credit
        else:
            balance = account.opening_balance + total_credit - total_debit
        
        t_account_data.append({
            'account': account,
            'debit_entries': debit_entries,
            'credit_entries': credit_entries,
            'total_debit': total_debit,
            'total_credit': total_credit,
        })
    
    # ----- STEP 5: Trial Balance -----
    trial_balance = []
    total_tb_debits = 0
    total_tb_credits = 0
    
    for account in accounts_with_entries:
        # Calculate balance
        total_debit = JournalEntry.objects.filter(account=account).aggregate(Sum('debit'))['debit__sum'] or 0
        total_credit = JournalEntry.objects.filter(account=account).aggregate(Sum('credit'))['credit__sum'] or 0
        
        if account.normal_balance == 'D':
            balance = account.opening_balance + total_debit - total_credit
        else:
            balance = account.opening_balance + total_credit - total_debit
        
        if balance != 0:
            if account.normal_balance == 'D':
                if balance > 0:
                    trial_balance.append({'account': account, 'debit': balance, 'credit': 0})
                    total_tb_debits += balance
                else:
                    trial_balance.append({'account': account, 'debit': 0, 'credit': abs(balance)})
                    total_tb_credits += abs(balance)
            else:
                if balance > 0:
                    trial_balance.append({'account': account, 'debit': 0, 'credit': balance})
                    total_tb_credits += balance
                else:
                    trial_balance.append({'account': account, 'debit': abs(balance), 'credit': 0})
                    total_tb_debits += abs(balance)
    
    # ----- STEP 6: Categorize accounts for display -----
    asset_accounts = [a for a in accounts_with_entries if a.account_type == 'asset']
    liability_accounts = [a for a in accounts_with_entries if a.account_type == 'liability']
    equity_accounts = [a for a in accounts_with_entries if a.account_type == 'equity']
    revenue_accounts = [a for a in accounts_with_entries if a.account_type == 'revenue']
    expense_accounts = [a for a in accounts_with_entries if a.account_type == 'expense']
    
    context = {
        # Summary cards
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'total_equity': total_equity,
        'total_revenue': total_revenue,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        
        # Journals
        'journals': journals,
        
        # Account lists
        'asset_accounts': asset_accounts,
        'liability_accounts': liability_accounts,
        'equity_accounts': equity_accounts,
        'revenue_accounts': revenue_accounts,
        'expense_accounts': expense_accounts,
        
        # Registers
        'assets_list': assets_list,
        'liabilities_list': liabilities_list,
        'equity_list': equity_list,
        
        # T-Accounts
        't_accounts': t_account_data,
        
        # Trial Balance
        'trial_balance': trial_balance,
        'total_tb_debits': total_tb_debits,
        'total_tb_credits': total_tb_credits,
        
        # Income Statement
        'total_revenue_amount': total_revenue,
        'total_expense_amount': total_expenses,
        'net_income': net_profit,
    }
    
    return render(request, 'laundry/accounting_full.html', context)
# ============================================================
# EXPORTS (PDF, EXCEL)
# ============================================================
@login_required
def export_trial_balance_pdf(request):
    from django.db.models import Sum
    
    accounts = LedgerAccount.objects.filter(is_active=True)
    trial_data = []
    total_debits = 0
    total_credits = 0
    
    for account in accounts:
        total_debit = JournalEntry.objects.filter(account=account).aggregate(Sum('debit'))['debit__sum'] or 0
        total_credit = JournalEntry.objects.filter(account=account).aggregate(Sum('credit'))['credit__sum'] or 0
        
        if account.normal_balance == 'D':
            net_balance = account.opening_balance + total_debit - total_credit
            if net_balance > 0:
                trial_data.append({'account': account, 'debit': net_balance, 'credit': 0})
                total_debits += net_balance
            else:
                trial_data.append({'account': account, 'debit': 0, 'credit': abs(net_balance)})
                total_credits += abs(net_balance)
        else:
            net_balance = account.opening_balance + total_credit - total_debit
            if net_balance > 0:
                trial_data.append({'account': account, 'debit': 0, 'credit': net_balance})
                total_credits += net_balance
            else:
                trial_data.append({'account': account, 'debit': abs(net_balance), 'credit': 0})
                total_debits += abs(net_balance)
    
    response = HttpResponse(content_type='text/plain')
    response.write("=" * 60 + "\n")
    response.write("TRIAL BALANCE\n")
    response.write("=" * 60 + "\n\n")
    response.write(f"Date: {timezone.now().date()}\n\n")
    response.write(f"{'Account':<35} {'Debit':>12} {'Credit':>12}\n")
    response.write("-" * 60 + "\n")
    
    for item in trial_data:
        response.write(f"{item['account'].name[:35]:<35} {item['debit']:>12.2f} {item['credit']:>12.2f}\n")
    
    response.write("-" * 60 + "\n")
    response.write(f"{'TOTAL':<35} {total_debits:>12.2f} {total_credits:>12.2f}\n")
    response.write("=" * 60 + "\n")
    
    if total_debits == total_credits:
        response.write("? TRIAL BALANCE IS BALANCED!\n")
    else:
        response.write(f"?? DIFFERENCE: {total_debits - total_credits:,.2f}\n")
    
    return response

@login_required
def export_trial_balance_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.db.models import Sum
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="trial_balance.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    headers = ['Account Code', 'Account Name', 'Account Type', 'Debit (KSh)', 'Credit (KSh)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    accounts = LedgerAccount.objects.filter(is_active=True)
    row = 2
    total_debits = 0
    total_credits = 0
    
    for account in accounts:
        total_debit = JournalEntry.objects.filter(account=account).aggregate(Sum('debit'))['debit__sum'] or 0
        total_credit = JournalEntry.objects.filter(account=account).aggregate(Sum('credit'))['credit__sum'] or 0
        
        if account.normal_balance == 'D':
            net_balance = account.opening_balance + total_debit - total_credit
            if net_balance > 0:
                ws.cell(row=row, column=1, value=account.account_code)
                ws.cell(row=row, column=2, value=account.name)
                ws.cell(row=row, column=3, value=account.get_account_type_display())
                ws.cell(row=row, column=4, value=float(net_balance))
                ws.cell(row=row, column=5, value=0)
                total_debits += net_balance
            else:
                ws.cell(row=row, column=1, value=account.account_code)
                ws.cell(row=row, column=2, value=account.name)
                ws.cell(row=row, column=3, value=account.get_account_type_display())
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=float(abs(net_balance)))
                total_credits += abs(net_balance)
        else:
            net_balance = account.opening_balance + total_credit - total_debit
            if net_balance > 0:
                ws.cell(row=row, column=1, value=account.account_code)
                ws.cell(row=row, column=2, value=account.name)
                ws.cell(row=row, column=3, value=account.get_account_type_display())
                ws.cell(row=row, column=4, value=0)
                ws.cell(row=row, column=5, value=float(net_balance))
                total_credits += net_balance
            else:
                ws.cell(row=row, column=1, value=account.account_code)
                ws.cell(row=row, column=2, value=account.name)
                ws.cell(row=row, column=3, value=account.get_account_type_display())
                ws.cell(row=row, column=4, value=float(abs(net_balance)))
                ws.cell(row=row, column=5, value=0)
                total_debits += abs(net_balance)
        row += 1
    
    ws.cell(row=row, column=3, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(total_debits)).font = Font(bold=True)
    ws.cell(row=row, column=5, value=float(total_credits)).font = Font(bold=True)
    
    for col in range(1, 6):
        ws.column_dimensions[chr(64 + col)].width = 20
    
    wb.save(response)
    return response

@login_required
def export_income_statement_pdf(request):
    response = HttpResponse(content_type='text/plain')
    response.write("=" * 60 + "\n")
    response.write("INCOME STATEMENT\n")
    response.write("=" * 60 + "\n\n")
    response.write(f"Date: {timezone.now().date()}\n\n")
    
    revenue_accounts = LedgerAccount.objects.filter(account_type='revenue', is_active=True)
    expense_accounts = LedgerAccount.objects.filter(account_type='expense', is_active=True)
    
    total_revenue = 0
    total_expenses = 0
    
    response.write("REVENUE:\n")
    response.write("-" * 40 + "\n")
    for acc in revenue_accounts:
        balance = acc.current_balance
        response.write(f"{acc.name:<35} {balance:>12.2f}\n")
        total_revenue += balance
    
    response.write("-" * 40 + "\n")
    response.write(f"{'Total Revenue':<35} {total_revenue:>12.2f}\n\n")
    
    response.write("EXPENSES:\n")
    response.write("-" * 40 + "\n")
    for acc in expense_accounts:
        balance = acc.current_balance
        response.write(f"{acc.name:<35} {balance:>12.2f}\n")
        total_expenses += balance
    
    response.write("-" * 40 + "\n")
    response.write(f"{'Total Expenses':<35} {total_expenses:>12.2f}\n\n")
    response.write("=" * 60 + "\n")
    response.write(f"Net Income: {total_revenue - total_expenses:>25.2f}\n")
    
    return response

@login_required
def export_balance_sheet_pdf(request):
    response = HttpResponse(content_type='text/plain')
    response.write("=" * 60 + "\n")
    response.write("BALANCE SHEET\n")
    response.write("=" * 60 + "\n\n")
    response.write(f"Date: {timezone.now().date()}\n\n")
    
    assets = LedgerAccount.objects.filter(account_type='asset', is_active=True)
    liabilities = LedgerAccount.objects.filter(account_type='liability', is_active=True)
    equity = LedgerAccount.objects.filter(account_type='equity', is_active=True)
    
    total_assets = 0
    total_liabilities = 0
    total_equity = 0
    
    response.write("ASSETS:\n")
    response.write("-" * 40 + "\n")
    for acc in assets:
        balance = acc.current_balance
        response.write(f"{acc.name:<35} {balance:>12.2f}\n")
        total_assets += balance
    
    response.write("-" * 40 + "\n")
    response.write(f"{'Total Assets':<35} {total_assets:>12.2f}\n\n")
    
    response.write("LIABILITIES:\n")
    response.write("-" * 40 + "\n")
    for acc in liabilities:
        balance = acc.current_balance
        response.write(f"{acc.name:<35} {balance:>12.2f}\n")
        total_liabilities += balance
    
    response.write("-" * 40 + "\n")
    response.write(f"{'Total Liabilities':<35} {total_liabilities:>12.2f}\n\n")
    
    response.write("EQUITY:\n")
    response.write("-" * 40 + "\n")
    for acc in equity:
        balance = acc.current_balance
        response.write(f"{acc.name:<35} {balance:>12.2f}\n")
        total_equity += balance
    
    response.write("-" * 40 + "\n")
    response.write(f"{'Total Equity':<35} {total_equity:>12.2f}\n\n")
    response.write("=" * 60 + "\n")
    response.write(f"Assets: {total_assets:>25.2f}\n")
    response.write(f"Liabilities + Equity: {total_liabilities + total_equity:>17.2f}\n")
    
    if total_assets == total_liabilities + total_equity:
        response.write("? BALANCE SHEET IS BALANCED!\n")
    else:
        response.write(f"?? DIFFERENCE: {total_assets - total_liabilities - total_equity:,.2f}\n")
    
    return response

@login_required
def export_balance_sheet_excel(request):
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="balance_sheet.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    row = 1
    
    ws.cell(row=row, column=1, value="ASSETS").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='27ae60', end_color='27ae60', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1
    
    assets = LedgerAccount.objects.filter(account_type='asset', is_active=True)
    total_assets = 0
    for acc in assets:
        ws.cell(row=row, column=1, value=acc.name)
        ws.cell(row=row, column=2, value=float(acc.current_balance))
        total_assets += acc.current_balance
        row += 1
    
    ws.cell(row=row, column=1, value="Total Assets").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(total_assets)).font = Font(bold=True)
    row += 2
    
    ws.cell(row=row, column=1, value="LIABILITIES").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='e74c3c', end_color='e74c3c', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1
    
    liabilities = LedgerAccount.objects.filter(account_type='liability', is_active=True)
    total_liabilities = 0
    for acc in liabilities:
        ws.cell(row=row, column=1, value=acc.name)
        ws.cell(row=row, column=2, value=float(acc.current_balance))
        total_liabilities += acc.current_balance
        row += 1
    
    ws.cell(row=row, column=1, value="Total Liabilities").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(total_liabilities)).font = Font(bold=True)
    row += 2
    
    ws.cell(row=row, column=1, value="EQUITY").font = Font(bold=True, size=14)
    ws.cell(row=row, column=1).fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    ws.cell(row=row, column=1).font = Font(bold=True, color='FFFFFF')
    row += 1
    
    equity = LedgerAccount.objects.filter(account_type='equity', is_active=True)
    total_equity = 0
    for acc in equity:
        ws.cell(row=row, column=1, value=acc.name)
        ws.cell(row=row, column=2, value=float(acc.current_balance))
        total_equity += acc.current_balance
        row += 1
    
    ws.cell(row=row, column=1, value="Total Equity").font = Font(bold=True)
    ws.cell(row=row, column=2, value=float(total_equity)).font = Font(bold=True)
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    wb.save(response)
    return response


# ============================================================
# API
# ============================================================
@login_required
def api_order_count(request):
    pending_count = Order.objects.filter(status__in=['received', 'washing', 'drying', 'folding']).count()
    return JsonResponse({'pending': pending_count})


# ============================================================
# ASSET MANAGEMENT
# ============================================================
@login_required
def asset_list(request):
    assets = Asset.objects.all().order_by('-created_at')
    total_assets_value = assets.aggregate(Sum('current_value'))['current_value__sum'] or 0
    # REMOVED: total_depreciation - no longer needed
    
    context = {
        'assets': assets,
        'total_assets_value': total_assets_value,
        # REMOVED: total_depreciation
    }
    return render(request, 'laundry/assets.html', context)

@login_required
def asset_create(request):
    if request.method == 'POST':
        try:
            asset = Asset.objects.create(
                asset_code=request.POST.get('asset_code'),
                name=request.POST.get('name'),
                asset_type=request.POST.get('asset_type'),
                purchase_date=request.POST.get('purchase_date'),
                purchase_price=request.POST.get('purchase_price'),
                current_value=request.POST.get('current_value'),
                location=request.POST.get('location', ''),
                description=request.POST.get('description', '')
            )
            return JsonResponse({'success': True, 'message': 'Asset added successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return render(request, 'laundry/asset_form.html')

@login_required
def asset_edit(request, pk):
    asset = get_object_or_404(Asset, pk=pk)
    if request.method == 'POST':
        try:
            asset.asset_code = request.POST.get('asset_code')
            asset.name = request.POST.get('name')
            asset.asset_type = request.POST.get('asset_type')
            asset.purchase_date = request.POST.get('purchase_date')
            asset.purchase_price = request.POST.get('purchase_price')
            asset.current_value = request.POST.get('current_value')
            asset.location = request.POST.get('location', '')
            asset.description = request.POST.get('description', '')
            asset.save()
            return JsonResponse({'success': True, 'message': 'Asset updated successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    
    return render(request, 'laundry/asset_form.html', {'asset': asset})
@login_required
def asset_delete(request, pk):
    if request.method == 'POST':
        try:
            asset = get_object_or_404(Asset, pk=pk)
            asset.delete()
            return JsonResponse({'success': True, 'message': 'Asset deleted successfully!'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Invalid request'})


# ============================================================
# SMART TRANSACTION SYSTEM
# ============================================================
@login_required
def smart_transaction(request):
    if request.method == 'POST':
        description = request.POST.get('description')
        amount = request.POST.get('amount')
        transaction_type = request.POST.get('transaction_type', 'purchase')
        
        analyzer = TransactionAnalyzer()
        result = analyzer.analyze(description, Decimal(amount), transaction_type)
        
        transaction = Transaction.objects.create(
            description=description,
            amount=amount,
            transaction_type=transaction_type,
            detected_category=result['detected_category'],
            confidence_score=result['confidence_score'],
            created_by=request.user
        )
        
        if result['confidence_score'] > 0.80:
            generator = AccountingEntryGenerator(transaction)
            journal = generator.generate()
            
            transaction.verified = True
            transaction.verified_by = request.user
            transaction.verified_at = timezone.now()
            transaction.detected_account = _get_account_for_category(result['detected_category'])
            transaction.related_journal = journal
            transaction.save()
            
            return JsonResponse({
                'success': True,
                'message': '? Transaction recorded and accounting entries created!',
                'transaction_id': transaction.id,
                'journal_id': journal.id,
                'result': result
            })
        else:
            return JsonResponse({
                'success': True,
                'requires_verification': True,
                'transaction_id': transaction.id,
                'result': result,
                'message': 'Please verify the transaction category before creating accounting entries.'
            })
    
    return render(request, 'laundry/smart_transaction.html')

@login_required
def confirm_transaction(request, transaction_id):
    if request.method == 'POST':
        try:
            transaction = Transaction.objects.get(id=transaction_id)
            transaction.verified = True
            transaction.verified_by = request.user
            transaction.verified_at = timezone.now()
            transaction.detected_account = _get_account_for_category(transaction.detected_category)
            
            generator = AccountingEntryGenerator(transaction)
            journal = generator.generate()
            transaction.related_journal = journal
            transaction.save()
            
            return JsonResponse({
                'success': True, 
                'message': 'Transaction confirmed and accounting entries created!',
                'redirect_url': '/accounting/'
            })
        except Transaction.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Transaction not found'}, status=404)
    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)

@login_required
def view_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    return render(request, 'laundry/transaction_detail.html', {'transaction': transaction})

@login_required
def edit_transaction(request, transaction_id):
    transaction = get_object_or_404(Transaction, id=transaction_id)
    if request.method == 'POST':
        transaction.description = request.POST.get('description')
        transaction.amount = request.POST.get('amount')
        transaction.transaction_type = request.POST.get('transaction_type')
        transaction.detected_category = request.POST.get('detected_category')
        transaction.save()
        return JsonResponse({'success': True, 'message': 'Transaction updated successfully!'})
    
    return render(request, 'laundry/transaction_edit.html', {'transaction': transaction})

@login_required
def transaction_list(request):
    transactions = Transaction.objects.all().order_by('-created_at')
    context = {
        'transactions': transactions,
        'total_transactions': transactions.count(),
        'total_amount': transactions.aggregate(Sum('amount'))['amount__sum'] or 0,
    }
    return render(request, 'laundry/transaction_list.html', context)


# ============================================================
# HELPER FUNCTIONS
# ============================================================
def _get_account_for_category(category):
    account_map = {
        'asset': '1000',
        'liability': '2000',
        'revenue': '4000',
        'expense': '5000',
        'equity': '3000',
    }
    account_code = account_map.get(category, '5000')
    try:
        return LedgerAccount.objects.get(account_code=account_code)
    except LedgerAccount.DoesNotExist:
        return None


# ============================================================
# AUTHENTICATION
# ============================================================
@login_required
def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('/login/')


# ============================================================
# USER MANAGEMENT (Owner Only)
# ============================================================
@login_required
@owner_required
def user_management(request):
    from django.contrib.auth.models import User
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'laundry/user_management.html', {'users': users})

@login_required
@owner_required
def user_create(request):
    from django.contrib.auth.models import User
    from django.contrib import messages
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username {username} already exists!')
            return redirect('user_management')
        
        user = User.objects.create_user(username=username, password=password, email=email)
        user.profile.role = 'staff'
        user.profile.phone = phone
        user.profile.is_active = True
        user.profile.save()
        
        messages.success(request, f'Staff user {username} created successfully!')
        return redirect('user_management')
    
    return render(request, 'laundry/user_form.html', {'title': 'Add Staff User'})

@login_required
@owner_required
def user_edit(request, pk):
    from django.contrib.auth.models import User
    from django.contrib import messages
    
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user.email = request.POST.get('email')
        user.profile.phone = request.POST.get('phone')
        
        new_password = request.POST.get('password')
        if new_password:
            user.set_password(new_password)
        
        user.profile.save()
        user.save()
        messages.success(request, f'User {user.username} updated!')
        return redirect('user_management')
    
    return render(request, 'laundry/user_form.html', {'user': user, 'title': 'Edit User'})

@login_required
@owner_required
def user_toggle(request, pk):
    from django.contrib import messages
    user = get_object_or_404(User, pk=pk)
    user.profile.is_active = not user.profile.is_active
    user.profile.save()
    status = 'activated' if user.profile.is_active else 'deactivated'
    messages.success(request, f'User {user.username} {status}!')
    return redirect('user_management')

@login_required
@owner_required
def user_delete(request, pk):
    from django.contrib import messages
    if request.method == 'POST':
        user = get_object_or_404(User, pk=pk)
        username = user.username
        
        if user == request.user:
            messages.error(request, 'You cannot delete your own account!')
        else:
            user.delete()
            messages.success(request, f'User {username} deleted successfully!')
        return redirect('user_management')
    return redirect('user_management')

@login_required
@owner_required
def user_management(request):
    from django.contrib.auth.models import User
    return render(request, 'laundry/users.html', {
        'users': User.objects.all(),
        'section': 'users',
        'page_title': 'User Management'
    })
